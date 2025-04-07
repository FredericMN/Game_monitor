# utils/crawler.py

import sys  # 确保导入sys模块
import os
import time
import random
import datetime
import openpyxl
import re
import shutil  # 用于copyfile
import threading
import requests

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor, as_completed
from PySide6.QtCore import Signal # Import Signal

# 导入任务管理器
from utils.task_manager import task_manager

# 导入WebDriverHelper
try:
    from utils.webdriver_helper import WebDriverHelper
except ImportError:
    # 如果导入失败，创建一个简易的辅助类
    class WebDriverHelper:
        @staticmethod
        def create_driver(options=None, headless=True, progress_callback=None):
            # 默认使用原始方法创建浏览器
            opt = options or webdriver.EdgeOptions()
            if headless:
                opt.add_argument("--headless")
            return webdriver.Edge(
                service=EdgeService(EdgeChromiumDriverManager().install()),
                options=opt
            )
        
        @staticmethod
        def quit_driver(driver):
            if driver:
                try:
                    driver.quit()
                except:
                    pass

MAX_WORKERS = 3

def progress_log_callback(callback, message):
    """统一日志输出，处理Qt Signal"""
    if callback:
        try:
            # Check if it's a Qt Signal instance
            if isinstance(callback, Signal):
                callback.emit(message)
            else:
                # Assume it's a regular callable
                callback(message)
        except RuntimeError as e:
            # Handle cases where the underlying Qt object might be destroyed
            print(f"[Callback Error] {e}: {message}") 
        except Exception as e:
             print(f"[Callback Exception] {type(e).__name__}: {message}")
    else:
        print(message) # Fallback if no callback provided

def random_delay(min_sec=0.5, max_sec=1.5):
    """避免爬取过快"""
    time.sleep(random.uniform(min_sec, max_sec))

def safe_execute(func):
    """装饰器统一处理异常"""
    def wrapper(*args, progress_callback=None, **kwargs):
        try:
            return func(*args, progress_callback=progress_callback, **kwargs)
        except Exception as e:
            progress_log_callback(progress_callback, f"执行出错: {str(e)}")
            if 'progress_percent_callback' in kwargs:
                kwargs['progress_percent_callback'](100, 0)
            return None
    return wrapper

# -----------------------------------------------------------------------------
# 新游爬虫
# -----------------------------------------------------------------------------
@safe_execute
def crawl_new_games(
    start_date_str,
    end_date_str,
    progress_callback=None,
    enable_version_match=True,
    progress_percent_callback=None
):
    """
    1. 无"序号"列；列头: [ "日期", "游戏名称", "状态", "厂商", "类型", "评分" ]
    2. 分天并发，每日爬完就输出并写到 Excel，但最终爬完后，再做一次"按日期升序"全表排序。
    3. 若 enable_version_match=True，则自动在同一个 Excel 里匹配版号并存储（不改名/不另存）。
    """

    # 提示
    progress_log_callback(progress_callback,
        "任务开始，可能存在网络缓慢等情况，请耐心等待。")

    try:
        sd = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        ed = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        if ed < sd:
            msg="结束日期不能早于起始日期，操作已取消。"
            progress_log_callback(progress_callback, msg)
            if progress_percent_callback:
                progress_percent_callback(100, 0)
            return
    except:
        msg="日期格式不正确，请使用 yyyy-MM-dd。操作已取消。"
        progress_log_callback(progress_callback, msg)
        if progress_percent_callback:
            progress_percent_callback(100, 0)
        return

    date_list=[]
    delta=(ed - sd).days+1
    for i in range(delta):
        date_list.append(sd + datetime.timedelta(days=i))

    total_dates=len(date_list)
    if total_dates==0:
        progress_log_callback(progress_callback,
            "无可爬取日期。")
        if progress_percent_callback:
            progress_percent_callback(100,0)
        return

    # 创建Excel
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    excel_filename = f"游戏数据_{today_str}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title="游戏数据"
    ws.append(["日期","游戏名称","状态","厂商","类型","评分"])  # 无"序号"
    wb.save(excel_filename)

    progress_log_callback(progress_callback,
        f"共需爬取 {total_dates} 天的新游信息，将分天爬取...")

    def crawl_one_day(d):
        """
        返回 (day_str, [ (name, status, man, types, rating) ])，若结构异常 => raise RuntimeError
        """
        day_str = d.strftime("%Y-%m-%d")
        opt = webdriver.EdgeOptions()
        # 创建进度回调的包装器
        wrapped_callback = lambda message, percent=None: helper_progress_callback(progress_callback, message)
        # 使用WebDriverHelper创建WebDriver
        driver = WebDriverHelper.create_driver(
            options=opt, 
            headless=True,
            progress_callback=wrapped_callback
        )
        # 注册WebDriver到任务管理器
        task_manager.register_webdriver(driver)
        
        results = []
        try:
            url = f"https://www.taptap.cn/app-calendar/{day_str}"
            driver.get(url)
            
            # 修改：先等待页面加载完成
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "div.daily-event-list__content"))
                )
            except:
                raise RuntimeError("网页结构疑似存在更新变动，请联系开发者进行解决。")

            # 查找游戏元素
            games = driver.find_elements(
                By.CSS_SELECTOR,
                "div.daily-event-list__content > a.tap-router"
            )
            
            # 如果没有游戏，直接返回空列表
            if not games:
                return (day_str, [])

            for g in games:
                random_delay()
                try:
                    name_el=g.find_element(
                        By.CSS_SELECTOR,"div.daily-event-app-info__title")
                    name=name_el.get_attribute("content").strip()
                except:
                    name="未知名称"
                try:
                    t_list=g.find_elements(
                        By.CSS_SELECTOR,
                        "div.daily-event-app-info__tag div.tap-label-tag")
                    types="/".join([t.text.strip() for t in t_list])
                except:
                    types="未知类型"
                try:
                    rating_el=g.find_element(
                        By.CSS_SELECTOR,
                        "div.daily-event-app-info__rating .tap-rating__number")
                    rating=rating_el.text.strip()
                except:
                    rating="未知评分"
                try:
                    st_el=g.find_element(
                        By.CSS_SELECTOR,"span.event-type-label__title")
                    status=st_el.text.strip()
                except:
                    try:
                        st2=g.find_element(
                            By.CSS_SELECTOR,"div.event-recommend-label__title")
                        status=st2.text.strip()
                    except:
                        status="未知状态"

                # 子页面
                href=g.get_attribute("href")
                if not href.startswith("http"):
                    href="https://www.taptap.cn"+href
                driver.execute_script("window.open(arguments[0]);", href)
                driver.switch_to.window(driver.window_handles[1])
                random_delay()

                man="未知厂商"
                try:
                    # 等待包含厂商/发行/开发信息的父容器加载
                    # 使用 'div.row-card.app-intro' 作为更可靠的等待目标
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.row-card.app-intro"))
                    )
                    
                    # 查找所有可能的厂商/发行/开发信息块
                    # 这些信息通常位于 <a> 标签内，包含一个标签名 (厂商/发行/开发) 和对应的值
                    # 选择器 'div.flex-center--y a.tap-router' 定位到这些链接
                    info_elements = driver.find_elements(By.CSS_SELECTOR, "div.flex-center--y a.tap-router")
                    
                    # 创建一个字典来存储找到的标签和对应的值
                    possible_publishers = {}
                    for elem in info_elements:
                        try:
                            # 尝试提取标签名 (如 "厂商", "开发", "发行")
                            # 标签名位于 class 包含 'gray-06' 和 'mr-6' 的 div 中
                            label_element = elem.find_element(By.CSS_SELECTOR, "div.gray-06.mr-6")
                            label = label_element.text.strip()
                            
                            # 尝试提取对应的值 (公司名称)
                            # 公司名位于 class 包含 'tap-text' 和 'tap-text__one-line' 的 div 中
                            value_element = elem.find_element(By.CSS_SELECTOR, "div.tap-text.tap-text__one-line")
                            value = value_element.text.strip()
                            
                            # 如果成功提取到标签和值，则存入字典
                            if label and value:
                                possible_publishers[label] = value
                                
                        except Exception:
                            # 如果在处理某个元素时出错（例如，结构不匹配），则忽略该元素，继续处理下一个
                            continue 
                            
                    # 应用优先级逻辑选择厂商名称
                    # 定义期望的标签及其优先级顺序
                    priority = ["厂商", "发行", "开发"]
                    for key in priority:
                        # 检查字典中是否存在当前优先级的标签
                        if key in possible_publishers:
                            # 如果找到，则使用对应的值作为厂商名称，并停止查找
                            man = possible_publishers[key]
                            break 

                except Exception as e:
                    # 如果在等待或查找元素的整体过程中发生异常 (例如超时)，
                    # 则保留默认值 "未知厂商"，可以选择性地记录日志
                    # progress_log_callback(progress_callback, f"提取厂商信息时出错: {e}")
                    pass # 保留默认值

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

                results.append( (name, status, man, types, rating) )
        finally:
            # 取消注册并关闭WebDriver
            task_manager.unregister_webdriver(driver)
            WebDriverHelper.quit_driver(driver)
        return (day_str, results)

    completed=0

    # 并发: 以天为粒度
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 注册线程池到任务管理器 
        task_manager.register_thread_pool(executor)
        try:
            future_map={ executor.submit(crawl_one_day, d): d for d in date_list}
            for future in as_completed(future_map):
                d = future_map[future]
                try:
                    day_str, day_data = future.result()
                except Exception as e:
                    progress_log_callback(progress_callback, str(e))
                    if progress_percent_callback:
                        progress_percent_callback(100,0)
                    return

                if day_data:
                    wb_cur=openpyxl.load_workbook(excel_filename)
                    ws_cur=wb_cur.active
                    block=[f"\n=== [日期 {day_str}, 共{len(day_data)} 款游戏] ==="]
                    for (nm, st, man, ty, rt) in day_data:
                        ws_cur.append([day_str,nm,st,man,ty,rt])
                        block.append(
                            f"  * {nm}\n"
                            f"    状态：{st}\n"
                            f"    厂商：{man}\n"
                            f"    类型：{ty}\n"
                            f"    评分：{rt}"
                        )
                    wb_cur.save(excel_filename)
                    text="\n".join(block)
                    progress_log_callback(progress_callback, text)
                else:
                    progress_log_callback(progress_callback,
                        f"=== [日期 {day_str}] 无游戏信息 ===")

                completed+=1
                if progress_percent_callback:
                    val=int(completed*100/total_dates)
                    progress_percent_callback(val,0)
        finally:
            # 取消注册线程池
            task_manager.unregister_thread_pool(executor)

    # 全部爬完后 => 对整个Excel按"日期"升序排序
    final_wb=openpyxl.load_workbook(excel_filename)
    final_ws=final_wb.active
    data_rows=list(final_ws.values)
    headers=data_rows[0]
    body=data_rows[1:]
    def parse_date(ds):
        try:
            return datetime.datetime.strptime(ds, "%Y-%m-%d")
        except:
            return datetime.datetime(1970,1,1)
    body.sort(key=lambda row: parse_date(row[0]))
    final_ws.delete_rows(1, final_ws.max_row)
    final_ws.append(headers)
    for row in body:
        final_ws.append(row)
    final_wb.save(excel_filename)

    progress_log_callback(progress_callback,
        f"新游数据已保存至 {excel_filename} (已按日期升序整理)")

    if progress_percent_callback:
        progress_percent_callback(100,0)

    # 若自动版号匹配
    if enable_version_match:
        if progress_percent_callback:
            progress_percent_callback(0,1)
        match_version_numbers(
            excel_filename,
            progress_callback=progress_callback,
            progress_percent_callback=progress_percent_callback,
            stage=1,
            create_new_file=False
        )

# -----------------------------------------------------------------------------
# 版号匹配
# -----------------------------------------------------------------------------
def match_version_numbers(
    excel_filename,
    progress_callback=None,
    progress_percent_callback=None,
    stage=1,
    create_new_file=True
):
    """
    - 如果 create_new_file=True => 基于原文件创建副本，并在副本上进行后续操作
    - 如果 create_new_file=False => 在同文件追加
    - 分段输出(每2~3行)
    - 需添加表头: [ "游戏名称", "出版单位", "运营单位", "文号", "出版物号", "版号获批时间", "游戏类型", "申报类别", "是否多个结果" ]
    """
    import shutil  # 用于复制文件

    progress_log_callback(progress_callback,
        "开始版号匹配，可能存在网络缓慢等情况，请耐心等待。")

    if not os.path.exists(excel_filename):
        progress_log_callback(progress_callback,
            f"文件 {excel_filename} 不存在，无法进行版号匹配。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    # 若要求另存为新文件，则先创建副本，并使用该副本进行操作
    if create_new_file:
        base, ext = os.path.splitext(excel_filename)
        new_fname = f"{base}-已匹配版号{ext}"
        try:
            shutil.copyfile(excel_filename, new_fname)
            progress_log_callback(progress_callback,
                f"已基于原文件创建副本 {new_fname}，开始在副本上匹配版号。")
            excel_filename = new_fname  # 后续操作使用新副本
        except Exception as copy_err:
            progress_log_callback(progress_callback,
                f"创建文件副本失败: {copy_err}")
            if progress_percent_callback:
                progress_percent_callback(100, stage)
            return

    try:
        wb = openpyxl.load_workbook(excel_filename)
    except Exception as e:
        progress_log_callback(progress_callback,
            f"无法加载Excel：{excel_filename}, 错误: {str(e)}")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    ws = wb.active
    new_headers = [
        "游戏名称","出版单位","运营单位","文号","出版物号",
        "版号获批时间","游戏类型","申报类别","是否多个结果"
    ]
    current_cols = [c.value for c in ws[1]]
    for nh in new_headers:
        if nh not in current_cols:
            ws.cell(row=1, column=len(current_cols)+1, value=nh)
            current_cols.append(nh)
    wb.save(excel_filename)

    # 提前获取各字段对应的列索引，避免后续重复查找
    field_column_map = {}
    for field in new_headers:
        if field in current_cols:
            field_column_map[field] = current_cols.index(field) + 1

    name_col = None
    for idx, c in enumerate(current_cols, start=1):
        if c in ["游戏名称","名称"]:
            name_col = idx
            break

    if name_col is None:
        progress_log_callback(progress_callback,
            "未找到'游戏名称'或'名称'列，跳过版号匹配。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    game_list = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        val = row[name_col - 1].value
        if val:
            game_list.append((row_idx, val))

    total_count = len(game_list)
    if total_count == 0:
        progress_log_callback(progress_callback,
            "没有需要匹配版号的条目。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    progress_log_callback(progress_callback,
        f"需要匹配 {total_count} 条游戏数据...")

    fields_for_match = [
        "游戏名称",   # info[0]
        "出版单位",   # info[1]
        "运营单位",   # info[2]
        "文号",       # info[3]
        "出版物号",   # info[4]
        "版号获批时间", # info[5]
        "游戏类型",   # info[6]
        "申报类别",   # info[7]
        "是否多个结果" # info[8]
    ]

    cache = {}
    def fetch_game_info(g_name):
        if g_name in cache:
            return cache[g_name]
        
        # 创建WebDriver时添加性能优化选项
        opt = webdriver.EdgeOptions()
        # 禁用不必要的功能以提高性能
        opt.add_argument('--disable-extensions')
        opt.add_argument('--no-sandbox')
        opt.add_argument('--disable-dev-shm-usage')
        
        driver = None
        res = None
        try:
            # 创建进度回调的包装器
            wrapped_callback = lambda message, percent=None: helper_progress_callback(progress_callback, message)
            # 使用WebDriverHelper创建WebDriver
            driver = WebDriverHelper.create_driver(
                options=opt,
                headless=True,
                progress_callback=wrapped_callback
            )
            # 注册WebDriver到任务管理器
            task_manager.register_webdriver(driver)
            
            # 设置页面加载超时
            driver.set_page_load_timeout(30)
            driver.set_script_timeout(30)
            
            qn = re.sub(r'（[^）]*）', '', g_name)
            qn = re.sub(r'\([^)]*\)', '', qn).strip()
            url = (f"https://www.nppa.gov.cn/bsfw/jggs/cxjg/index.html?"
                   f"mc={qn}&cbdw=&yydw=&wh=undefined&description=#")
            
            # 添加重试机制
            max_retries = 3
            for retry in range(max_retries):
                try:
                    driver.get(url)
                    break
                except Exception as e:
                    if retry == max_retries - 1:
                        raise
                    time.sleep(2)  # 等待一段时间后重试
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#dataCenter"))
                )
            except Exception as e:
                progress_log_callback(progress_callback, 
                    f"游戏 {g_name} 网页加载超时或结构变动: {str(e)}")
                return None

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#dataCenter tr"))
                )
            except:
                rows2 = []
            else:
                rows2 = driver.find_elements(By.CSS_SELECTOR, "#dataCenter tr")

            if not rows2:
                res = None
            else:
                multiple_flag = "是" if len(rows2) > 1 else "否"
                exact_elem = None
                for rr in rows2:
                    time.sleep(0.3)
                    try:
                        t = rr.find_element(By.CSS_SELECTOR, "a").text.strip()
                        if t == g_name:
                            exact_elem = rr
                            break
                    except:
                        pass
                if not exact_elem:
                    exact_elem = rows2[0]
                res = extract_game_info(exact_elem, multiple_flag, driver)
        except requests.exceptions.RequestException as req_err:
            progress_log_callback(progress_callback, 
                f"游戏 {g_name} 网络请求异常: {str(req_err)}")
            res = None
        except Exception as e:
            progress_log_callback(progress_callback, 
                f"游戏 {g_name} 处理异常: {str(e)}")
            res = None
        finally:
            if driver:
                # 取消注册WebDriver 
                task_manager.unregister_webdriver(driver)
                WebDriverHelper.quit_driver(driver)

        cache[g_name] = res
        return res

    def extract_game_info(elem, multi_flag, driver):
        try:
            tds = elem.find_elements(By.TAG_NAME, "td")
            if len(tds) < 7:
                return None
            gn = tds[1].text.strip()
            pub = tds[2].text.strip()
            op = tds[3].text.strip()
            appr = tds[4].text.strip()
            pubn = tds[5].text.strip()
            ds = tds[6].text.strip()

            detail_url = ""
            try:
                detail_url = tds[1].find_element(By.TAG_NAME, "a").get_attribute("href")
            except:
                pass

            gtype, appcat = "", ""
            if detail_url:
                try:
                    driver.execute_script("window.open(arguments[0]);", detail_url)
                    driver.switch_to.window(driver.window_handles[1])
                    random_delay()
                    
                    # 添加等待以确保页面加载
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, ".cFrame.nFrame table tr"))
                        )
                    except:
                        pass  # 如果等待超时，继续处理已加载的内容
                    
                    lines = driver.find_elements(By.CSS_SELECTOR, ".cFrame.nFrame table tr")
                    for ln in lines:
                        try:
                            lab = ln.find_element(By.XPATH, "./td[1]").text.strip()
                            val = ln.find_element(By.XPATH, "./td[2]").text.strip()
                            if lab == "游戏类型":
                                gtype = val
                            elif lab == "申报类别":
                                appcat = val
                        except:
                            pass
                finally:
                    # 确保窗口关闭和切换回主窗口
                    try:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    except:
                        pass  # 忽略窗口操作可能出现的异常

            return [gn, pub, op, appr, pubn, ds, gtype, appcat, multi_flag]
        except Exception as e:
            progress_log_callback(progress_callback, 
                f"提取游戏详情异常: {str(e)}")
            return None

    partial_flush_size = 3
    completed = 0
    buffered = []
    results_map = {}
    
    # 使用线程锁保护Excel写入
    excel_lock = threading.Lock()

    def flush_results(buf):
        if not buf:
            return []
        buf.sort(key=lambda x: x[0])  # row_idx升序
        lines = []
        
        # 使用线程锁保护Excel文件写入
        with excel_lock:
            try:
                wb2 = openpyxl.load_workbook(excel_filename)
                ws2 = wb2.active
                
                for (r_idx, g_n, info) in buf:
                    if info:
                        # info顺序: [gn, pub, op, appr, pubn, ds, gtype, appcat, multi_flag]
                        block = [
                            f"--- [行 {r_idx}, 游戏名: {g_n}] ---",
                            f"  游戏名称：{info[0]}",
                            f"  出版单位：{info[1]}",
                            f"  运营单位：{info[2]}",
                            f"  文号：{info[3]}",
                            f"  出版物号：{info[4]}",
                            f"  版号获批时间：{info[5]}",
                            f"  游戏类型：{info[6]}",
                            f"  申报类别：{info[7]}",
                            f"  是否多个结果：{info[8]}"
                        ]
                        lines.extend(block)
                        
                        # 使用预先计算的列索引映射写入数据
                        for i, field_name in enumerate(fields_for_match):
                            if field_name in field_column_map:
                                col_idx = field_column_map[field_name]
                                ws2.cell(row=r_idx, column=col_idx, value=info[i])
                    else:
                        lines.append(f"--- [行 {r_idx}, 游戏名: {g_n}] => 未查询到该游戏版号信息。")

                wb2.save(excel_filename)
            except Exception as e:
                lines.append(f"写入Excel出错: {str(e)}")
        
        text = "\n".join(lines)
        progress_log_callback(progress_callback, text)
        return []

    # 调整并发数量，避免过多线程导致资源争用
    max_workers = min(MAX_WORKERS, 2)  # 版号匹配时限制并发数
    
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        # 注册线程池到任务管理器
        task_manager.register_thread_pool(ex)
        try:
            future_map = {}
            for (r_idx, g_n) in game_list:
                future = ex.submit(fetch_game_info, g_n)
                future_map[future] = (r_idx, g_n)

            for future in as_completed(future_map):
                row_i, g_na = future_map[future]
                try:
                    info = future.result()
                except Exception as e:
                    progress_log_callback(progress_callback, f"处理游戏 {g_na} 时出错: {str(e)}")
                    info = None
                
                results_map[row_i] = info
                buffered.append((row_i, g_na, info))
                completed += 1
                
                if len(buffered) >= partial_flush_size:
                    buffered = flush_results(buffered)
                
                if progress_percent_callback:
                    pr = int(completed * 100 / total_count)
                    progress_percent_callback(pr, stage)
        finally:
            # 取消注册线程池 
            task_manager.unregister_thread_pool(ex)

    if buffered:
        buffered = flush_results(buffered)

    # 若 create_new_file=True，已在开头复制并操作副本，不再另存
    if create_new_file:
        pass
    else:
        progress_log_callback(progress_callback,
            f"版号匹配完成，结果已追加到原文件 {excel_filename}")

    if progress_percent_callback:
        progress_percent_callback(100, stage)

def helper_progress_callback(callback, message):
    """处理WebDriverHelper的进度回调，过滤和简化WebDriver输出信息"""
    # 定义需要保留的关键信息关键词
    key_messages = [
        "使用已缓存的WebDriver", 
        "未找到缓存", 
        "正在下载", 
        "下载/安装完成",
        "创建浏览器实例",
        "浏览器实例创建成功",
        "缓存WebDriver配置",
        "创建WebDriver时出错"
    ]
    
    # 检查消息是否包含关键信息
    show_message = False
    for key in key_messages:
        if key in message:
            show_message = True
            break
    
    # 简化消息，删除过多细节
    if "缓存目录" in message or "缓存文件" in message:
        return  # 忽略缓存路径信息
    
    if "检测到Edge浏览器版本" in message:
        return  # 忽略版本检测信息
    
    # 只显示关键节点信息
    if show_message and callback:
        # 进一步简化消息内容
        if "使用已缓存的WebDriver" in message:
            simplified = "使用已缓存的WebDriver配置"
        elif "未找到缓存的WebDriver" in message:
            simplified = "未找到可用WebDriver缓存，需要下载配置"
        elif "正在下载" in message:
            simplified = "正在下载WebDriver"
        elif "下载/安装完成" in message:
            simplified = "WebDriver下载完成"
        elif "创建浏览器实例" in message and "成功" not in message:
            simplified = "正在启动浏览器"
        elif "浏览器实例创建成功" in message:
            simplified = "浏览器启动成功"
        elif "缓存WebDriver配置" in message:
            simplified = "WebDriver配置已缓存，下次将加速启动"
        elif "创建WebDriver时出错" in message:
            simplified = f"启动浏览器失败: {message.split(':', 1)[1] if ':' in message else ''}"
        else:
            simplified = message
        
        progress_log_callback(callback, simplified)
